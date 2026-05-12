import pdfplumber
import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
import openpyxl
import os
import streamlit as st
import re

from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
import io

'''
A função pdf_to_excel lê o ficheiro pdf e converte-o para um ficheiro excel
As páginas de pdf que são convertidas para excel são aquelas que contém tabelas com medidas
'''

def is_vazio(x):
    return x is None or (isinstance(x, str) and x.strip() in ("", "None"))

def limpar_linhas_vazias(df):
    """
    Remove elementos vazios a partir da 4ª coluna (índice 3) e desloca o conteúdo para trás.
    As primeiras 4 colunas mantêm-se intactas.
    """
    def limpar_linha(row):
        # Manter as primeiras 4 colunas (índices 0, 1, 2,3)
        primeiras_4 = list(row[:4])
        
        # Filtrar apenas elementos que não são vazios a partir da coluna 5
        elementos_validos_resto = [x for x in row[4:] if not is_vazio(x)]
        
        # Preencher com strings vazias até ao comprimento original do resto
        while len(elementos_validos_resto) < len(row[4:]):
            elementos_validos_resto.append("")
        
        # Combinar as 4 primeiras colunas com o resto processado
        return primeiras_4 + elementos_validos_resto
    
    # Aplicar a função a cada linha
    df_limpo = df.apply(limpar_linha, axis=1, result_type='expand')
    return df_limpo

def extract_sections_from_text(text,cliente):
    split1  = text.split("Ref:")
    if len(split1) != 1:
        ref = split1[1].split("ANGLOTEX - CONFECÇÕES, LDA.")[0]
    else:
        ref = "ND"
    split2 = text.split("LDA.")
    if len(split2) != 1:
        name = split2[1].split(cliente)[0]
    else:
        name  = "ND"
    ref = re.sub(r'[\\/*?:"<>|]', "_", ref)
    if len(ref)>30:
        ref = ref[:20]
    return ref.strip(), name.strip()

def pdf_to_excel(nome_pdf,excel_name,cliente):
   with pdfplumber.open(nome_pdf) as pdf:
        # Lista para guardar todas as tabelas de todas as páginas
        todas_tabelas = []

        for i, page in enumerate(pdf.pages):
            tables = page.extract_tables({
                "vertical_strategy": "lines",
                "horizontal_strategy": "lines"
            })
            if i == 0:
                text = page.extract_text()
                ref_text, name_text = extract_sections_from_text(text,cliente)
            for table in tables:
                df = pd.DataFrame(table).astype(str)
                todas_tabelas.append(df)

        # Juntar todas as tabelas num único DataFrame
        if todas_tabelas:
            final_df = pd.DataFrame()
            for df in todas_tabelas:
                final_df = pd.concat([final_df, df, pd.DataFrame([[""] * len(df.columns)])], ignore_index=True)

            # Procurar a linha onde a primeira coluna diz ....
            mask = final_df.iloc[:, 0].str.strip().str.lower() == "bordados e estampados"
            mask2 = final_df.iloc[:, 0].str.strip().str.lower() == "acessorios"
            mask3 = final_df.iloc[:, 0].str.strip().str.lower() == "malhas e tecidos"
            mask4 = final_df.iloc[:, 0].str.strip().str.lower() == "ponto de control"
            mask5 = final_df.iloc[:, 0].str.strip().str.lower() == "acabamentos a peça"

            if all(m.any() for m in (mask2, mask3, mask4)):
                idx2 = mask2.idxmax()  # índice da primeira ocorrência de acessorios
                idx3 = mask3.idxmax()  # índice da primeira ocorrência de malhas e tecidos
                idx4 = mask4.idxmax()  # índice da primeira ocorrência de ponto de control

                """
                inf malhas e tecidos
                """
                page_2_df = final_df.iloc[idx3+1:idx4, :]
                #escolher apenas linhas que tenham UN ou MT ou KG na coluna D (índice 3)
                col_unidade = page_2_df.iloc[:, 3].str.strip().str.lower()
                mask_unidades = (col_unidade == 'un')| (col_unidade == 'kg') | (col_unidade == 'mt')
                page_2_df = page_2_df[mask_unidades].reset_index(drop=True)
                # Identificar valores none
                page_2_df = limpar_linhas_vazias(page_2_df)

                """
                inf tabela ponto de controlo
                """
                page_3_df = final_df.iloc[idx4+1:, :]
                #escolher apenas linhas que tenham apenas as operações até desconto
                col_opere = page_3_df.iloc[:, 0].str.strip().str.lower()
                mask_opere = (col_opere == 'acessorios')| (col_opere == 'malhas e tecidos') | (col_opere == 'malha tinturaria') | (col_opere == 'corte') | (col_opere == 'bord./est. (animações)') | (col_opere == 'confecção') | (col_opere == 'embalamento') | (col_opere == 'linhas') | (col_opere == 'desconto') | (col_opere == 'acabamentos a peça') | (col_opere == 'gastos gerais') | (col_opere == 'transporte')| (col_opere == 'margem corte') | (col_opere == 'comissão')| (col_opere == 'margem')
                page_3_df = page_3_df[mask_opere].reset_index(drop=True)

                # Identificar colunas totalmente vazias
                colunas_para_remover3 = [col for col in page_3_df.columns if all(is_vazio(x) for x in page_3_df[col])]
                # Remover colunas vazias
                page_3_df = page_3_df.drop(columns=colunas_para_remover3)

                """
                inf acessorios
                """
                page_5_df = final_df.iloc[idx2+1:idx3, :]
                # Identificar valores none
                page_5_df = limpar_linhas_vazias(page_5_df)


                if mask.any() and mask5.any():
                    idx = mask.idxmax()  # índice da primeira ocorrência de bordados e estampados
                    idx5 = mask5.idxmax()   # índice da primeira ocorrência de acabamentos a peça

                    """
                    inf bordados e estampados
                    """
                    page_1_df = final_df.iloc[idx+1:idx2, :]
                    # Identificar valores none
                    page_1_df = limpar_linhas_vazias(page_1_df)


                    """
                    inf acabementos à peça
                    """
                    page_4_df = final_df.iloc[idx5+1:idx, :]
                    # Identificar valores none
                    page_4_df = limpar_linhas_vazias(page_4_df)

                elif mask5.any() and not mask.any():
                    idx5 = mask5.idxmax()   # índice da primeira ocorrência de acabamentos a peça

                    """
                    inf bordados e estampados
                    """
                    page_1_df = pd.DataFrame()  # DataFrame vazio para bordados e estampados, já que não existe essa secção

                    """
                    inf acabementos à peça
                    """
                    page_4_df = final_df.iloc[idx5+1:idx2, :]
                    # Identificar valores none
                    page_4_df = limpar_linhas_vazias(page_4_df)
                
                elif mask.any() and not mask5.any():
                    idx = mask.idxmax()  # índice da primeira ocorrência de bordados e estampados

                    """
                    inf bordados e estampados
                    """
                    page_1_df = final_df.iloc[idx+1:idx2, :]
                    # Identificar valores none
                    page_1_df = limpar_linhas_vazias(page_1_df)

                    """
                    inf acabementos à peça
                    """
                    page_4_df = pd.DataFrame()  # DataFrame vazio para acabamentos à peça, já que não existe essa secção
                elif not mask.any() and not mask5.any():
                    #se não tiver bordados nem acabamentos à peça
                    """
                    inf bordados e estampados
                    """
                    page_1_df = pd.DataFrame()
                    """
                    inf acabementos à peça
                    """
                    page_4_df = pd.DataFrame()  # DataFrame vazio para acabamentos à peça, já que não existe essa secção

            else:
                missing = [name for name, m in (("malhas e tecidos", mask3),
                                                ("acessorios", mask2),
                                                ("ponto de control", mask4)) if not m.any()]
                raise ValueError(f"Faltam secções obrigatórias: {', '.join(missing)}")

            # Escrever no Excel
            with pd.ExcelWriter(excel_name, engine='xlsxwriter') as writer:
                page_1_df.to_excel(writer, sheet_name='Page_1', index=False, header=False)
                page_2_df.to_excel(writer, sheet_name='Page_2', index=False, header=False)
                page_3_df.to_excel(writer, sheet_name='Page_3', index=False, header=False)
                page_4_df.to_excel(writer, sheet_name='Page_4', index=False, header=False)
                page_5_df.to_excel(writer, sheet_name='Page_5', index=False, header=False)
    

        return ref_text, name_text
   

def add_images(pdf_path,excel_path,inf_texto):
    # Carregar o arquivo Excel existente
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Inserir linhas vazias no topo (para espaço das imagens)
    ws.insert_rows(1, 5)

        # Adicionar texto no topo
    i=1
    for texto in inf_texto:
        cell=ws.cell(row=i, column=2, value=texto)  # Insere o texto na coluna A
        cell.font = Font(bold=True, size=14) 
        i+=1

    image_paths=[]
    with pdfplumber.open(pdf_path) as pdf:
        row, col = 1, 3  # Linha inicial e coluna inicial

        for i, page in enumerate(pdf.pages):
            if i == 0:
                images = page.images  # Obtém as imagens da página
                for j, img in enumerate(images):
                    # Extrair a imagem
                    img_data = img["stream"].get_data()
                    image = Image.open(io.BytesIO(img_data))

                    max_height_px = 100  # altura máxima em pixels

                    # Calcular factor de escala para respeitar a altura máxima
                    scale = min(1.0, max_height_px / image.height)
                    new_width = int(image.width * scale)
                    new_height = int(image.height * scale)

                    image = image.resize((new_width, new_height))

                    # Salvar como arquivo temporário
                    img_path = f"temp_img_{i}_{j}.png"
                    image.save(img_path)
                    image_paths.append(img_path)    

                    # Adicionar ao Excel
                    excel_img = ExcelImage(img_path)
                    
                    # Também garantir dimensões no objecto openpyxl (opcional)
                    excel_img.width = new_width
                    excel_img.height = new_height
                    ws.add_image(excel_img, f"{openpyxl.utils.get_column_letter(col)}{row}")

    # Inserir linhas vazias no topo (para espaço das imagens)
    excel_img_logo = ExcelImage("Logo anglotex PRETO.png")
    # Também garantir dimensões no objecto openpyxl (opcional)
    excel_img_logo.width = 100
    excel_img_logo.height = 80
    ws.add_image(excel_img_logo,f"{openpyxl.utils.get_column_letter(1)}{1}")

    # Salvar o arquivo Excel atualizado
    wb.save(excel_path)
    # Remover os ficheiros das imagens após inserir no Excel
    for img_path in image_paths:
        os.remove(img_path)
