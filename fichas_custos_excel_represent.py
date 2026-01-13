import pdfplumber
import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
import openpyxl
import os

from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
import io

'''
A função pdf_to_excel lê o ficheiro pdf e converte-o para um ficheiro excel
As páginas de pdf que são convertidas para excel são aquelas que contém tabelas com medidas
'''

def is_vazio(x):
    return x is None or (isinstance(x, str) and x.strip() in ("", "None"))

def limpar_linhas_vazias(df):
    """
    Remove elementos vazios a partir da 4ª coluna (índice 3) e desloca o conteúdo para trás.
    As primeiras 4 colunas mantêm-se intactas.
    """
    def limpar_linha(row):
        # Manter as primeiras 4 colunas (índices 0, 1, 2,3)
        primeiras_4 = list(row[:4])
        
        # Filtrar apenas elementos que não são vazios a partir da coluna 5
        elementos_validos_resto = [x for x in row[4:] if not is_vazio(x)]
        
        # Preencher com strings vazias até ao comprimento original do resto
        while len(elementos_validos_resto) < len(row[4:]):
            elementos_validos_resto.append("")
        
        # Combinar as 4 primeiras colunas com o resto processado
        return primeiras_4 + elementos_validos_resto
    
    # Aplicar a função a cada linha
    df_limpo = df.apply(limpar_linha, axis=1, result_type='expand')
    return df_limpo

def extract_sections_from_text(text):
    split1  = text.split("Ref:")
    if len(split1) != 1:
        ref = split1[1].split("ANGLOTEX - CONFECÇÕES, LDA.")[0]
    else:
        ref = "referencia não encontrada, por favor adicionar manualmente"
    split2 = text.split("ANGLOTEX - CONFECÇÕES, LDA.")
    if len(split2) != 1:
        name = split2[1].split("Matéria")[0]
    else:
        name  = "modelo não encontrado, por favor adicionar manualmente"
    return ref.strip(), name.strip()

def pdf_to_excel(nome_pdf,excel_name):
   with pdfplumber.open(nome_pdf) as pdf:
        # Lista para guardar todas as tabelas de todas as páginas
        todas_tabelas = []

        for i, page in enumerate(pdf.pages):
            tables = page.extract_tables({
                "vertical_strategy": "lines",
                "horizontal_strategy": "lines"
            })
            if i == 0:
                text = page.extract_text()
                ref_text, name_text = extract_sections_from_text(text)
            for table in tables:
                df = pd.DataFrame(table).astype(str)
                todas_tabelas.append(df)

        # Juntar todas as tabelas num único DataFrame
        if todas_tabelas:
            final_df = pd.DataFrame()
            for df in todas_tabelas:
                final_df = pd.concat([final_df, df, pd.DataFrame([[""] * len(df.columns)])], ignore_index=True)

            # Procurar a linha onde a primeira coluna diz "Ponto de controlo"
            mask = final_df.iloc[:, 0].str.strip().str.lower() == "bordados e estampados"
            mask2 = final_df.iloc[:, 0].str.strip().str.lower() == "acessorios"
            mask3 = final_df.iloc[:, 0].str.strip().str.lower() == "malhas e tecidos"
            mask4 = final_df.iloc[:, 0].str.strip().str.lower() == "ponto de control"
            mask5 = final_df.iloc[:, 0].str.strip().str.lower() == "acabamentos a peça"
            if all(m.any() for m in (mask, mask2, mask3, mask4)):
                idx = mask.idxmax()  # índice da primeira ocorrência de bordados e estampados
                idx2 = mask2.idxmax()  # índice da primeira ocorrência de acessorios
                idx3 = mask3.idxmax()  # índice da primeira ocorrência de malhas e tecidos
                idx4 = mask4.idxmax()  # índice da primeira ocorrência de ponto de control

                page_1_df = final_df.iloc[idx+1:idx2, :]
                # Identificar valores none
                page_1_df = limpar_linhas_vazias(page_1_df)

                page_2_df = final_df.iloc[idx3+1:idx4, :]
                #escolher apenas linhas que tenham UN ou MT ou KG na coluna D (índice 3)
                col_unidade = page_2_df.iloc[:, 3].str.strip().str.lower()
                mask_unidades = (col_unidade == 'un')| (col_unidade == 'kg') | (col_unidade == 'mt')
                page_2_df = page_2_df[mask_unidades].reset_index(drop=True)
                # Identificar valores none
                page_2_df = limpar_linhas_vazias(page_2_df)

                page_3_df = final_df.iloc[idx4+1:, :]
                #escolher apenas linhas que tenham apenas as operações até desconto
                col_opere = page_3_df.iloc[:, 0].str.strip().str.lower()
                mask_opere = (col_opere == 'acessorios')| (col_opere == 'malhas e tecidos') | (col_opere == 'malha tinturaria') | (col_opere == 'corte') | (col_opere == 'bord./est. (animações)') | (col_opere == 'confecção') | (col_opere == 'embalamento') | (col_opere == 'linhas') | (col_opere == 'desconto') | (col_opere == 'acabamentos a peça') | (col_opere == 'gastos gerais') | (col_opere == 'transporte')| (col_opere == 'margem corte') | (col_opere == 'comissão')| (col_opere == 'margem')
                page_3_df = page_3_df[mask_opere].reset_index(drop=True)

                # Identificar colunas totalmente vazias
                colunas_para_remover3 = [col for col in page_3_df.columns if all(is_vazio(x) for x in page_3_df[col])]
                # Remover colunas vazias
                page_3_df = page_3_df.drop(columns=colunas_para_remover3)

            else:
                missing = [name for name, m in (("bordados e estampados", mask),
                                                ("acessorios", mask2),
                                                ("malhas e tecidos", mask3),
                                                ("ponto de control", mask4)) if not m.any()]
                raise ValueError(f"Faltam secções obrigatórias: {', '.join(missing)}")
            
            if mask5.any():
                idx5 = mask5.idxmax()  # índice da primeira ocorrência de acabamentos a peça
                page_4_df = final_df.iloc[idx5+1:idx, :]
                # Identificar valores none
                page_4_df = limpar_linhas_vazias(page_4_df)


            # Escrever no Excel
            with pd.ExcelWriter(excel_name, engine='xlsxwriter') as writer:
                page_1_df.to_excel(writer, sheet_name='Page_1', index=False, header=False)
                page_2_df.to_excel(writer, sheet_name='Page_2', index=False, header=False)
                page_3_df.to_excel(writer, sheet_name='Page_3', index=False, header=False)
                if mask5.any():
                    page_4_df.to_excel(writer, sheet_name='Page_4', index=False, header=False)
    

        return ref_text, name_text


# ...existing code...
def trim_excel_before_marker(excel_path,excel_saida):

    """
    preparar excel
    """
    linhas_excel=[]
    #header_colunas = ["","","consumption per garmet","", "cost per garmet (€)", "total cost per garmet (€)"]
    header_colunas = ["","", "total cost per garmet (€)"]
    #linhas_excel.append(header_colunas)

    sheets = pd.read_excel(excel_path, sheet_name=None, header=None, dtype=str)
    #devolve algo do género {'Sheet1': df1, 'Sheet2': df2, …}


    """
    Obter valor da margem a aplicar - está na sheet 3 do excel
    """

    #procurar na segunda sheet do excel
    second_sheet_name = list(sheets.keys())[2]
    #garante que todos os valores são string e sem NaN (substitui NaN pela string "")
    df2 = sheets[second_sheet_name].fillna('').astype(str)

    #selecionar a primeira coluna
    first_col2 = df2.iloc[:, 0].str.strip().str.lower()
    
    # margem_aplicar
    marker_margem_aplicar = "margem"
    mask_margem_aplicar = first_col2.eq(marker_margem_aplicar.lower())
    if mask_margem_aplicar.any():
        margem_aplicar_idxs = mask_margem_aplicar[mask_margem_aplicar].index.tolist()
        margem_aplicar_idx = margem_aplicar_idxs[-1]  # Pega no último índice
    else:
        raise ValueError(f"'{marker_margem_aplicar}' não encontrado na segunda folha do excel.")

    perc = pd.to_numeric(df2.iloc[margem_aplicar_idx, 1], errors='coerce')

    percent_value = perc/100


    """
    Calculo do consumo e custo por malha - SHEET 2 DO EXCEL
    """
    sheets = pd.read_excel(excel_path, sheet_name=None, header=None, dtype=str)
    #devolve algo do género {'Sheet1': df1, 'Sheet2': df2, …}

    #escolher a segunda sheet do excel
    first_sheet_name = list(sheets.keys())[1]
    #garante que todos os valores são string e sem NaN (substitui NaN pela string "")
    df = sheets[first_sheet_name].fillna('').astype(str)

    #verificar quantas malhas diferentes existem
    first_col = df.iloc[:, 0].str.strip().str.lower()
    malhas_indices = first_col[first_col != ""].index.tolist()
    ultima_linha = len(df) - 1 #ultima linha da pagina de excel
    num_malhas = len(malhas_indices)

    """
    FIM DE - Obter o consumo e preço por malha 

    """

    """
    CALCULAR RESTANTES COMPONENTES DO PREÇO: ARTWORKS, WASHING, CMT(CORTE,CONFEÇÃO,EMBALAMENTO), OTHER COSTS, ACESSÓRIOS
    """

    #df2 e first_col2 já estão definidos anteriormente para a sheet 3 do excel

    # acessorios
    marker_acessorios = "Acessorios"
    mask_acessorios = first_col2.eq(marker_acessorios.lower())
    if not mask_acessorios.any():
        acessorios_cost=0
        perc_acessorios = 0
    else:
        acessorios_idx = mask_acessorios.idxmax()
        acessorios_cost = pd.to_numeric(df2.iloc[acessorios_idx, 2], errors='coerce')
        perc_acessorios = acessorios_cost * percent_value

    #margem corte
    marker_margem_corte = "Margem Corte"
    mask_margem_corte = first_col2.eq(marker_margem_corte.lower())
    if not mask_margem_corte.any(): 
        margem_corte_cost=0
    else:
        margem_corte_idx = mask_margem_corte.idxmax()
        margem_corte_cost = pd.to_numeric(df2.iloc[margem_corte_idx, 2], errors='coerce')
        margem_corte_cost = margem_corte_cost * (1+percent_value)

    #valor que será dividido por Malhas, CMT, Artworks e Washing
    add_cost_div= margem_corte_cost + perc_acessorios
    

    # comissão (fecho)
    marker_comissao = "Comissão"
    mask_comissao = first_col2.eq(marker_comissao.lower())
    if not mask_comissao.any():
        comissao_cost=0
    else:
        comissao_idx = mask_comissao.idxmax()
        comissao_cost = pd.to_numeric(df2.iloc[comissao_idx, 2], errors='coerce')

    acessories_final_cost = acessorios_cost + comissao_cost


    # para contar quantos componentes terei que dividir a margem do corte e percentagem dos acessorios, para além de CMT, malhas e artworks
    count = 3


    # CMT (MANIFACTURING COST)
    markers_cmt = ["corte", "confecção", "embalamento", "linhas"]
    indices_cmt = []
    for marker in markers_cmt:
        mask_cmt = first_col2.eq(marker.lower())
        if not mask_cmt.any():
            markers_cmt.remove(marker)
            print(f"Atenção: '{marker}' não encontrado na tabela de Ponto de Control. Não será considerado no custo CMT")
            continue
            #raise ValueError(f"'{marker}' não encontrado na segunda folha do excel.")
        cmt_idx = mask_cmt.idxmax()
        indices_cmt.append(cmt_idx)

    cmt_cost = 0
    for idx in indices_cmt:
        cost = pd.to_numeric(df2.iloc[idx, 2], errors='coerce')
        if not pd.isna(cost):
            cmt_cost += cost
    cmt_margem_cost = cmt_cost * (1+percent_value)


    # artworks
    marker_descontos = "Desconto"
    mask_descontos = first_col2.eq(marker_descontos.lower())
    if not mask_descontos.any():
        descontos_cost=0
    else:
        descontos_idx = mask_descontos.idxmax()
        descontos_cost = pd.to_numeric(df2.iloc[descontos_idx, 2], errors='coerce')
    
    marker_artworks = "Bord./Est. (Animações)"
    mask_artworks = first_col2.eq(marker_artworks.lower())
    if not mask_artworks.any():
        artworks_cost=0
    else:
        #primeira sheet que contem os artworks descriminados
        primeiro_sheet_name = list(sheets.keys())[0]
        #garante que todos os valores são string e sem NaN (substitui NaN pela string "")
        df3 = sheets[primeiro_sheet_name].fillna('').astype(str)
        desconte_add = descontos_cost / len(df3)


    # washing
    marker_washing = "Acabamentos a Peça"
    mask_washing = first_col2.eq(marker_washing.lower())
    if not mask_washing.any():  
        washing_cost=0
    else:
        #primeira sheet que contem os artworks descriminados
        sheet4_name = list(sheets.keys())[3]
        #garante que todos os valores são string e sem NaN (substitui NaN pela string "")
        df4 = sheets[sheet4_name].fillna('').astype(str)
        count +=1

    div_value = add_cost_div / count

    #adicionar informação das malhas
    for i in range(0,num_malhas):
        div_value_per_malha = div_value/num_malhas
        if i < num_malhas-1:
            linha_inf = []
            linha_inf.append(df.iloc[malhas_indices[i],0])  # codigo da malha
            linha_inf.append(f"{df.iloc[malhas_indices[i],1]}")  # artigo da malha
            #consumo=pd.to_numeric(df.iloc[malhas_indices[i], -3], errors='coerce')
            #linha_inf.append(consumo)  # consumo da malha
            #linha_inf.append(df.iloc[malhas_indices[i],3]) #unidade da malha
            soma1=pd.to_numeric(df.iloc[malhas_indices[i]:malhas_indices[i+1], -1], errors='coerce').sum() 
            #linha_inf.append(soma1)  # preço antes de aplicar a margem
            linha_inf.append(round(float(soma1*(1+percent_value)+div_value_per_malha),2))  # preço após aplicar a margem e soma da parte dividida
            linhas_excel.append(linha_inf)
        else:
            linha_inf = []
            linha_inf.append(df.iloc[malhas_indices[i],0])  # codigo da malha
            linha_inf.append(f"{df.iloc[malhas_indices[i],1]}")  # artigo da malha
            #consumo=pd.to_numeric(df.iloc[malhas_indices[i], -3], errors='coerce')
            #linha_inf.append(consumo)  # consumo
            #linha_inf.append(df.iloc[malhas_indices[i],3]) #unidade da malha
            soma1=pd.to_numeric(df.iloc[malhas_indices[i]:ultima_linha+1, -1], errors='coerce').sum()
            #linha_inf.append(soma1)  # preço antes de aplicar a margem
            linha_inf.append(round(float(soma1*(1+percent_value)+div_value_per_malha),2))  # preço após aplicar a margem e soma da parte dividida
            linhas_excel.append(linha_inf)

    #adicionar informação cmt
    #linhas_excel.append(["","CMT", "", "", cmt_cost, cmt_margem_cost + div_value])
    linhas_excel.append(["","CMT", round(float(cmt_margem_cost + div_value),2)])

    #adicionar informação acessorios
    #linhas_excel.append(["","Trims", "", "", acessorios_cost, acessories_final_cost])
    linhas_excel.append(["","Trims", round(float(acessories_final_cost),2)])
    #adicionar informação artworks
    qtd_adicionar  = (div_value/len(df3))+desconte_add
    for i in range(0,len(df3)):
        linha_inf = []
        linha_inf.append(df3.iloc[i,0])  # codigo do artwork
        linha_inf.append(df3.iloc[i,1])  # artigo do artwork
        #linha_inf.append("") # consumo do artwork
        #linha_inf.append("") # unidade do artwork
        custo_artwork = pd.to_numeric(df3.iloc[i, -1], errors='coerce')
        #linha_inf.append(custo_artwork)  # preço antes de aplicar a margem
        linha_inf.append(round(float(custo_artwork*(1+percent_value)) + qtd_adicionar,2))  # preço após aplicar a margem e soma da parte dividida 
        linhas_excel.append(linha_inf)

    #adicionar informação washing caso exista
    if mask_washing.any():  
        qtd_adicionar_washing  = (div_value/len(df4))
        for i in range(0,len(df4)):
            linha_inf = []
            linha_inf.append(df4.iloc[i,0])  # codigo do washing
            linha_inf.append(df4.iloc[i,1])  # artigo do washing
            #linha_inf.append("") # consumo do washing
            #linha_inf.append("") # unidade do washing
            custo_washing = pd.to_numeric(df4.iloc[i, -1], errors='coerce')
            #linha_inf.append(custo_washing)  # preço antes de aplicar a margem
            linha_inf.append(round(float(custo_washing*(1+percent_value)) + qtd_adicionar_washing,2))  # preço após aplicar a margem e soma da parte dividida 
            linhas_excel.append(linha_inf)


    # other costs
    marker_other_costs = ['Gastos Gerais', 'Transporte']
    indices_other_costs = []
    for marker in marker_other_costs:
        mask_other_costs = first_col2.eq(marker.lower())
        if not mask_other_costs.any():
            raise ValueError(f"'{marker}' não encontrado na segunda folha do excel.")
        other_costs_idx = mask_other_costs.idxmax()
        indices_other_costs.append(other_costs_idx)
    
    other_costs = 0
    for idx in indices_other_costs:
        cost = pd.to_numeric(df2.iloc[idx, 2], errors='coerce')
        if not pd.isna(cost):
            other_costs += cost
    
    other_costs_final = other_costs * (1+percent_value)

    #linhas_excel.append(["","Other", "", "", other_costs,other_costs_final]) 
    linhas_excel.append(["","Other",round(float(other_costs_final),2)])

    """
    FIM - CALCULAR RESTANTES COMPONENTES DO PREÇO: ARTWORKS, WASHING, CMT(CORTE,CONFEÇÃO,EMBALAMENTO), OTHER COSTS, ACESSÓRIOS
    
    variaveis a usar: cmt_cost, acessorios_cost, artworks_cost, washing_cost
    """
    
    """
    CRIAR TABELA COM TODAS AS INFORMAÇÕES
    """

    df_final = pd.DataFrame(linhas_excel,columns = header_colunas)

    total_cost_per_garment = df_final["total cost per garmet (€)"].sum()
    #nova_linha = pd.DataFrame([["", "Total per garment", "", "", "", total_cost_per_garment]], columns=header_colunas)
    nova_linha = pd.DataFrame([["", "Total per garment", total_cost_per_garment]], columns=header_colunas)
    df_final = pd.concat([df_final, nova_linha], ignore_index=True)
          # Escrever no Excel
    with pd.ExcelWriter(excel_saida, engine='openpyxl') as writer:
        df_final.to_excel(writer, sheet_name='Sheet1', index=False,header=True)

    
    # Formatar
    wb = load_workbook(excel_saida)
    #com workbook os indices começam em 1
    ws = wb.active
    
    for sheet in wb.worksheets:
        last_row = len(sheet['A'])  # número de linhas na coluna A
        # Alterar a altura de todas as linhas
        for row in sheet.iter_rows():
            row[0].parent.row_dimensions[row[0].row].height = 18.5
        
        # Alterar a altura da quinta linha que é onde ficará o header depois de adicionar a imagem e referencia
        sheet.row_dimensions[6].height = 50 # Linha 1 com altura 25

        # Definir a largura da primeira coluna (A)
        sheet.column_dimensions['A'].width = 16
        # Definir a largura da primeira coluna (B)
        sheet.column_dimensions['B'].width = 56
        # Definir a largura da primeira coluna (C)
        sheet.column_dimensions['C'].width = 20



        #start_col_idx = openpyxl.utils.column_index_from_string('E')

        # Iterar por todas as colunas a partir de 'C' e ajustar a largura
        #for col_idx in range(start_col_idx, sheet.max_column + 1):
        #    column_letter = openpyxl.utils.get_column_letter(col_idx)
        #    sheet.column_dimensions[column_letter].width = 20

        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(size=12)  # Tamanho da letra 14
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='top')


        # Definir a espessura da borda
        border_style = Border(
            left=Side(style='thin', color='000000'),  # Bordas à esquerda
            right=Side(style='thin', color='000000'),  # Bordas à direita
            top=Side(style='thin', color='000000'),  # Bordas em cima
            bottom=Side(style='thin', color='000000')  # Bordas embaixo
        )

        # Iterar sobre todas as linhas e colunas da planilha
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border_style  # Definir a borda da célula

        # Tornar a última linha  em negrito
        for cell in sheet[last_row]:
            cell.font = Font(size=14, bold=True)


    wb.save(excel_saida)

    """
    FIM- CRIAR TABELA COM TODAS AS INFORMAÇÕES
    """ 

    return 


def add_images(pdf_path,excel_path,inf_texto):
    # Carregar o arquivo Excel existente
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Inserir linhas vazias no topo (para espaço das imagens)
    ws.insert_rows(1, 5)

        # Adicionar texto no topo
    i=1
    for texto in inf_texto:
        cell=ws.cell(row=i, column=1, value=texto)  # Insere o texto na coluna A
        cell.font = Font(bold=True, size=14) 
        i+=1

    image_paths=[]
    with pdfplumber.open(pdf_path) as pdf:
        row, col = 1, 3  # Linha inicial e coluna inicial

        for i, page in enumerate(pdf.pages):
            if i == 0:
                images = page.images  # Obtém as imagens da página
                for j, img in enumerate(images):
                    # Extrair a imagem
                    img_data = img["stream"].get_data()
                    image = Image.open(io.BytesIO(img_data))

                    max_height_px = 100  # altura máxima em pixels

                    # Calcular factor de escala para respeitar a altura máxima
                    scale = min(1.0, max_height_px / image.height)
                    new_width = int(image.width * scale)
                    new_height = int(image.height * scale)

                    image = image.resize((new_width, new_height))

                    # Salvar como arquivo temporário
                    img_path = f"temp_img_{i}_{j}.png"
                    image.save(img_path)
                    image_paths.append(img_path)    

                    # Adicionar ao Excel
                    excel_img = ExcelImage(img_path)
                    
                    # Também garantir dimensões no objecto openpyxl (opcional)
                    excel_img.width = new_width
                    excel_img.height = new_height
                    ws.add_image(excel_img, f"{openpyxl.utils.get_column_letter(col)}{row}")

    # Salvar o arquivo Excel atualizado
    wb.save(excel_path)
    # Remover os ficheiros das imagens após inserir no Excel
    for img_path in image_paths:
        os.remove(img_path)



