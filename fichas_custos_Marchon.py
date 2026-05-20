import pdfplumber
import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
import openpyxl
import os
import streamlit as st

from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
import io


def trim_excel_before_marker(excel_path,excel_saida):

    """
    preparar excel
    """
    linhas_excel=[]
    header_colunas = ["","", "total cost per garmet (€)"]


    sheets = pd.read_excel(excel_path, sheet_name=None, header=None, dtype=str)
    #devolve algo do género {'Sheet1': df1, 'Sheet2': df2, …}


    """
    Obter valor da margem e da comissão e aplicar - está na Page_3 do excel
    """
    
    third_sheet_name = list(sheets.keys())[2]
    #garante que todos os valores são string e sem NaN (substitui NaN pela string "")
    df2 = sheets[third_sheet_name].fillna('').astype(str)
    
    #selecionar a primeira coluna
    first_col2 = df2.iloc[:, 0].str.strip().str.lower()
    
    # margem (percentagem)
    marker_margem_aplicar = "margem"
    mask_margem_aplicar = first_col2.eq(marker_margem_aplicar.lower())
    if mask_margem_aplicar.any():
        margem_aplicar_idxs = mask_margem_aplicar[mask_margem_aplicar].index.tolist()
        margem_aplicar_idx = margem_aplicar_idxs[-1]  # Pega no último índice
        perc = pd.to_numeric(df2.iloc[margem_aplicar_idx, 1], errors='coerce')
        percent_value = perc/100
    else:
        raise ValueError(f"'{marker_margem_aplicar}' não encontrado na segunda folha do excel.")

    #comissão - a MARCHON não tem Comissão

    """
    FIM - Obter valor da margem e da comissão e aplicar - está na Page_3 do excel
    """

    #df2 e first_col2 já estão definidos anteriormente para a Page_3 do excel
    """
    CALCULAR VALOR A DISTRIBUIR
    """
    # para contar quantos componentes terei que DISTRIBUIR por Acessorios, CMT, malhas, artworks, washing
    count = 0
        
    #margem corte
    marker_margem_corte = "Margem Corte"
    mask_margem_corte = first_col2.eq(marker_margem_corte.lower())
    if not mask_margem_corte.any(): 
        margem_corte_cost=0
    else:
        margem_corte_idx = mask_margem_corte.idxmax()
        margem_corte_cost = pd.to_numeric(df2.iloc[margem_corte_idx, 2], errors='coerce')
        margem_corte_cost = margem_corte_cost*(1+percent_value)

    #valor que será dividido por Malhas, CMT, Artworks e Washing
    add_cost_div= margem_corte_cost

    """
    FIM - CALCULAR VALOR A DISTRIBUIR
    """

    # acessorios
    marker_acessorios = "Acessorios"
    mask_acessorios = first_col2.eq(marker_acessorios.lower())
    if mask_acessorios.any():
        count += 1
    
    """
    CALCULAR CMT
    """

    # CMT (MANIFACTURING COST)
    markers_cmt = ["corte", "confecção", "embalamento", "linhas"]
    indices_cmt = []
    for marker in markers_cmt:
        mask_cmt = first_col2.eq(marker.lower())
        if not mask_cmt.any():
            markers_cmt.remove(marker)
            print(f"Atenção: '{marker}' não encontrado na tabela de Ponto de Control. Não será considerado no custo CMT")
            continue
            #raise ValueError(f"'{marker}' não encontrado na segunda folha do excel.")
        cmt_idx = mask_cmt.idxmax()
        indices_cmt.append(cmt_idx)

    cmt_cost = 0
    for idx in indices_cmt:
        cost = pd.to_numeric(df2.iloc[idx, 2], errors='coerce')
        if not pd.isna(cost):
            cmt_cost += cost
    cmt_margem_cost = cmt_cost*(1+percent_value)

    #considerar CMT e Malhas
    count += 2

    # artworks
    marker_descontos = "Desconto"
    mask_descontos = first_col2.eq(marker_descontos.lower())
    if not mask_descontos.any():
        descontos_cost=0
    else:
        descontos_idx = mask_descontos.idxmax()
        descontos_cost = pd.to_numeric(df2.iloc[descontos_idx, 2], errors='coerce')
    
    marker_artworks = "Bord./Est. (Animações)"
    mask_artworks = first_col2.eq(marker_artworks.lower())
    if not mask_artworks.any():
        artworks_cost=0
    else:
        count +=1
        #primeira sheet que contem os artworks descriminados
        primeiro_sheet_name = list(sheets.keys())[0]
        #garante que todos os valores são string e sem NaN (substitui NaN pela string "")
        df3 = sheets[primeiro_sheet_name].fillna('').astype(str)
        #valor do desconto que será adicionado a cada artwork
        desconte_add = descontos_cost / len(df3)

    # washing
    marker_washing = "Acabamentos a Peça"
    mask_washing = first_col2.eq(marker_washing.lower())
    if not mask_washing.any():  
        washing_cost=0
    else:
        #primeira sheet que contem os artworks descriminados
        sheet4_name = list(sheets.keys())[3]
        #garante que todos os valores são string e sem NaN (substitui NaN pela string "")
        df4 = sheets[sheet4_name].fillna('').astype(str)
        count +=1

    #valor distribuir por categoria
    div_value = add_cost_div / count

    """
    Calculo do consumo e custo por malha - Page_2 DO EXCEL
    """
    sheets = pd.read_excel(excel_path, sheet_name=None, header=None, dtype=str)
    #devolve algo do género {'Sheet1': df1, 'Sheet2': df2, …}

    #escolher a segunda sheet do excel
    second_sheet_name = list(sheets.keys())[1]
    #garante que todos os valores são string e sem NaN (substitui NaN pela string "")
    df = sheets[second_sheet_name].fillna('').astype(str)

    #verificar quantas malhas diferentes existem
    first_col = df.iloc[:, 0].str.strip().str.lower()
    malhas_indices = first_col[first_col != ""].index.tolist()
    ultima_linha = len(df) - 1 #ultima linha da pagina de excel
    num_malhas = len(malhas_indices)

    #adicionar informação das malhas
    for i in range(0,num_malhas):
        #valor a adicionar a cada uma das malhas
        div_value_per_malha = div_value/num_malhas
        if i < num_malhas-1:
            linha_inf = []
            linha_inf.append(df.iloc[malhas_indices[i],0])  # codigo da malha
            linha_inf.append(f"{df.iloc[malhas_indices[i],1]}")  # artigo da malha
            soma_malha=pd.to_numeric(df.iloc[malhas_indices[i]:malhas_indices[i+1], -1], errors='coerce').sum()
            soma_malha_margem = soma_malha*(1+percent_value)
            linha_inf.append(round(float(soma_malha_margem+div_value_per_malha),2))  # preço após aplicar a margem e soma da parte dividida
            linhas_excel.append(linha_inf)
        else:
            linha_inf = []
            linha_inf.append(df.iloc[malhas_indices[i],0])  # codigo da malha
            linha_inf.append(f"{df.iloc[malhas_indices[i],1]}")  # artigo da malha
            soma_malha=pd.to_numeric(df.iloc[malhas_indices[i]:ultima_linha+1, -1], errors='coerce').sum()
            soma_malha_margem = soma_malha*(1+percent_value)
            linha_inf.append(round(float(soma_malha_margem+div_value_per_malha),2))# preço após aplicar a margem e soma da parte dividida
            linhas_excel.append(linha_inf)

    """
    ADICIONAR INFORMAÇÃO
    """
    #adicionar informação cmt
    linhas_excel.append(["","CMT", round(float(cmt_margem_cost + div_value),2)])

    #adicionar informação acessorios
    if mask_acessorios.any():
        #sheet que contem os acessorios descriminados
        sheet5_name = list(sheets.keys())[4]
        #garante que todos os valores são string e sem NaN (substitui NaN pela string "")
        df5 = sheets[sheet5_name].fillna('').astype(str)
        
        qtd_adicionar_acessorio = div_value/len(df5)
        for i in range(0,len(df5)):
            linha_inf=[]
            linha_inf.append(df5.iloc[i,0]) #codigo acessorio
            linha_inf.append(df5.iloc[i,1]) #descritivo do acessorio
            custo_acessorio = pd.to_numeric(df5.iloc[i, -1], errors='coerce')
            custo_acessorio_margem = custo_acessorio*(1+percent_value)
            linha_inf.append(round(float(custo_acessorio_margem + qtd_adicionar_acessorio),2))
            linhas_excel.append(linha_inf)

    #adicionar informação artworks
    #o valor a adicionar aos artworks será div_value mais o desconto-(dividido pela quantidade de artworks)
    if mask_artworks.any():
        qtd_adicionar  = (div_value/len(df3))+desconte_add
        for i in range(0,len(df3)):
            linha_inf = []
            linha_inf.append(df3.iloc[i,0])  # codigo do artwork
            linha_inf.append(df3.iloc[i,1])  # artigo do artwork
            custo_artwork = pd.to_numeric(df3.iloc[i, -1], errors='coerce')
            custo_artwork_margem = custo_artwork*(1+percent_value)
            linha_inf.append(round(float(custo_artwork_margem+ qtd_adicionar),2))  # preço após aplicar a margem e soma da parte dividida 
            linhas_excel.append(linha_inf)

    #adicionar informação washing caso exista
    if mask_washing.any():  
        qtd_adicionar_washing  = (div_value/len(df4))
        for i in range(0,len(df4)):
            linha_inf = []
            linha_inf.append(df4.iloc[i,0])  # codigo do washing
            linha_inf.append(df4.iloc[i,1])  # artigo do washing
            custo_washing = pd.to_numeric(df4.iloc[i, -1], errors='coerce')
            custo_washing_margem = custo_washing*(1+percent_value)
            linha_inf.append(round(float(custo_washing_margem + qtd_adicionar_washing),2))  # preço após aplicar a margem e soma da parte dividida 
            linhas_excel.append(linha_inf)


    # other costs
    marker_other_costs = ['Gastos Gerais', 'Transporte']
    indices_other_costs = []
    for marker in marker_other_costs:
        mask_other_costs = first_col2.eq(marker.lower())
        if not mask_other_costs.any():
            raise ValueError(f"'{marker}' não encontrado na segunda folha do excel.")
        other_costs_idx = mask_other_costs.idxmax()
        indices_other_costs.append(other_costs_idx)
    
    other_costs = 0
    for idx in indices_other_costs:
        cost = pd.to_numeric(df2.iloc[idx, 2], errors='coerce')
        if not pd.isna(cost):
            other_costs += cost
            
    other_costs_final = other_costs*(1+percent_value)

    linhas_excel.append(["","Other",round(float(other_costs_final),2)])

    """
    FIM - CALCULAR RESTANTES COMPONENTES DO PREÇO: ARTWORKS, WASHING, CMT(CORTE,CONFEÇÃO,EMBALAMENTO), OTHER COSTS, ACESSÓRIOS
    
    variaveis a usar: cmt_cost, acessorios_cost, artworks_cost, washing_cost
    """
    
    """
    CRIAR TABELA COM TODAS AS INFORMAÇÕES
    """

    df_final = pd.DataFrame(linhas_excel,columns = header_colunas)

    #total_cost_per_garment = df_final["total cost per garmet (€)"].sum()
    nr_df_final = 6+len(df_final)
    total_cost_per_garment = f"=SUM(C7:C{nr_df_final})"
    nova_linha = pd.DataFrame([["", "Total per garment", total_cost_per_garment]], columns=header_colunas)
    df_final = pd.concat([df_final, nova_linha], ignore_index=True)
          # Escrever no Excel
    with pd.ExcelWriter(excel_saida, engine='openpyxl') as writer:
        df_final.to_excel(writer, sheet_name='Sheet1', index=False,header=True)

    
    # Formatar
    wb = load_workbook(excel_saida)
    #com workbook os indices começam em 1
    ws = wb.active
    
    for sheet in wb.worksheets:
        last_row = len(sheet['A'])  # número de linhas na coluna A
        # Alterar a altura de todas as linhas
        for row in sheet.iter_rows():
            row[0].parent.row_dimensions[row[0].row].height = 18.5
        
        # Alterar a altura da quinta linha que é onde ficará o header depois de adicionar a imagem e referencia
        sheet.row_dimensions[6].height = 50 # Linha 1 com altura 25

        # Definir a largura da primeira coluna (A)
        sheet.column_dimensions['A'].width = 16
        # Definir a largura da primeira coluna (B)
        sheet.column_dimensions['B'].width = 56
        # Definir a largura da primeira coluna (C)
        sheet.column_dimensions['C'].width = 20



        #start_col_idx = openpyxl.utils.column_index_from_string('E')

        # Iterar por todas as colunas a partir de 'C' e ajustar a largura
        #for col_idx in range(start_col_idx, sheet.max_column + 1):
        #    column_letter = openpyxl.utils.get_column_letter(col_idx)
        #    sheet.column_dimensions[column_letter].width = 20

        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(size=12)  # Tamanho da letra 14
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='top')
                #if cell.column >= start_col_idx:  # Apenas células a partir da coluna 'D'
                #    cell.alignment = Alignment(horizontal='center', vertical='center')

        # Definir a espessura da borda
        border_style = Border(
            left=Side(style='thin', color='000000'),  # Bordas à esquerda
            right=Side(style='thin', color='000000'),  # Bordas à direita
            top=Side(style='thin', color='000000'),  # Bordas em cima
            bottom=Side(style='thin', color='000000')  # Bordas embaixo
        )

        # Iterar sobre todas as linhas e colunas da planilha
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border_style  # Definir a borda da célula

        # Tornar a última linha  em negrito
        for cell in sheet[last_row]:
            cell.font = Font(size=14, bold=True)


    wb.save(excel_saida)

    """
    FIM- CRIAR TABELA COM TODAS AS INFORMAÇÕES
    """ 

    return 





